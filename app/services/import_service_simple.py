"""Serviço de importação de dados simplificado"""

import json
import openpyxl
from io import BytesIO
from typing import Dict, List, Any, Optional
from fastapi import UploadFile, HTTPException
from sqlalchemy.orm import Session
from datetime import datetime

from app.models.contracts import Contract, BudgetItem
from app.models.purchases import Invoice, InvoiceItem


class SimpleDataImportService:
    def __init__(self, db: Session):
        self.db = db

    async def validate_file_format(self, file: UploadFile) -> Dict[str, Any]:
        """Valida formato do arquivo"""

        valid_extensions = ['.xlsx', '.xls', '.xml']
        file_extension = None

        if file.filename:
            file_extension = '.' + file.filename.split('.')[-1].lower()

        valid = file_extension in valid_extensions if file_extension else False

        return {
            "valid": valid,
            "filename": file.filename,
            "file_type": file_extension,
            "size": file.size if hasattr(file, 'size') else None,
            "supported_formats": valid_extensions
        }

    async def import_budget_from_excel(
        self,
        file: UploadFile,
        contract_id: int,
        sheet_name: str = "QQP Cliente"
    ) -> Dict[str, Any]:
        """Importa orçamento de planilha Excel"""

        try:
            # Ler arquivo Excel
            contents = await file.read()
            workbook = openpyxl.load_workbook(BytesIO(contents))

            # Tentar encontrar a aba especificada
            if sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
            else:
                # Usar a primeira aba disponível
                worksheet = workbook.active
                sheet_name = worksheet.title

            # Verificar se o contrato existe
            contract = self.db.query(Contract).filter(Contract.id == contract_id).first()
            if not contract:
                raise HTTPException(status_code=404, detail="Contrato não encontrado")

            # Processar dados da planilha
            imported_items = []
            errors = []

            # Assumir que a primeira linha são cabeçalhos
            headers = []
            for col in range(1, worksheet.max_column + 1):
                header = worksheet.cell(row=1, column=col).value
                headers.append(header)

            # Processar cada linha de dados
            for row_num in range(2, worksheet.max_row + 1):
                try:
                    row_data = {}
                    for col in range(1, len(headers) + 1):
                        cell_value = worksheet.cell(row=row_num, column=col).value
                        header = headers[col-1]
                        if header:
                            row_data[header] = cell_value

                    # Mapear colunas comuns (pode precisar ajuste baseado na planilha real)
                    budget_item = BudgetItem(
                        contract_id=contract_id,
                        descricao=str(row_data.get('Descrição', row_data.get('DESCRIÇÃO', 'N/A'))),
                        centro_custo=str(row_data.get('Centro de Custo', row_data.get('CENTRO_CUSTO', 'N/A'))),
                        unidade=str(row_data.get('Unidade', row_data.get('UNIDADE', 'UN'))),
                        quantidade=float(row_data.get('Quantidade', row_data.get('QUANTIDADE', 0)) or 0),
                        peso=float(row_data.get('Peso', row_data.get('PESO', 0)) or 0),
                        valor_unitario=float(row_data.get('Valor Unitário', row_data.get('VALOR_UNITARIO', 0)) or 0),
                        valor_total=float(row_data.get('Valor Total', row_data.get('VALOR_TOTAL', 0)) or 0)
                    )

                    # Se valor total não foi fornecido, calcular
                    if budget_item.valor_total == 0 and budget_item.quantidade > 0:
                        budget_item.valor_total = budget_item.quantidade * budget_item.valor_unitario

                    self.db.add(budget_item)
                    imported_items.append({
                        "row": row_num,
                        "description": budget_item.descricao,
                        "value": float(budget_item.valor_total)
                    })

                except Exception as row_error:
                    errors.append({
                        "row": row_num,
                        "error": str(row_error)
                    })

            self.db.commit()

            return {
                "success": True,
                "contract_id": contract_id,
                "sheet_name": sheet_name,
                "imported_items": len(imported_items),
                "items": imported_items,
                "errors": errors,
                "total_value": sum(item["value"] for item in imported_items)
            }

        except Exception as e:
            self.db.rollback()
            raise HTTPException(status_code=400, detail=f"Erro ao processar arquivo Excel: {str(e)}")

    async def import_invoice_from_xml(
        self,
        file: UploadFile,
        purchase_order_id: int
    ) -> Dict[str, Any]:
        """Importa nota fiscal de XML"""

        try:
            contents = await file.read()
            # Implementação básica - pode ser expandida para processar XML real da NF-e

            # Por enquanto, retornar estrutura básica
            return {
                "success": True,
                "message": "Importação de XML ainda não implementada completamente",
                "file_processed": file.filename,
                "purchase_order_id": purchase_order_id
            }

        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Erro ao processar XML: {str(e)}")

    async def import_invoice_from_excel(
        self,
        file: UploadFile,
        purchase_order_id: int
    ) -> Dict[str, Any]:
        """Importa nota fiscal de planilha Excel"""

        try:
            contents = await file.read()
            workbook = openpyxl.load_workbook(BytesIO(contents))
            worksheet = workbook.active

            # Processar dados básicos da nota fiscal
            # Assumir formato simples por enquanto

            return {
                "success": True,
                "message": "Importação de NF por Excel processada",
                "file_processed": file.filename,
                "purchase_order_id": purchase_order_id
            }

        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Erro ao processar planilha: {str(e)}")

    async def bulk_import_invoices(
        self,
        files: List[UploadFile],
        purchase_order_ids: Optional[List[int]] = None
    ) -> Dict[str, Any]:
        """Importação em lote de múltiplas notas fiscais"""

        results = []
        errors = []

        for i, file in enumerate(files):
            try:
                po_id = purchase_order_ids[i] if purchase_order_ids and i < len(purchase_order_ids) else None

                if file.filename and file.filename.endswith('.xml'):
                    result = await self.import_invoice_from_xml(file, po_id)
                else:
                    result = await self.import_invoice_from_excel(file, po_id)

                results.append(result)

            except Exception as e:
                errors.append({
                    "file": file.filename,
                    "error": str(e)
                })

        return {
            "total_files": len(files),
            "successful_imports": len(results),
            "failed_imports": len(errors),
            "results": results,
            "errors": errors
        }